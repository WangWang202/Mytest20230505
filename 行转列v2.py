import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.drawing.image import Image

input_file = 'C:/Users/Administrator/Desktop/抖音热销商品排行榜20230511.xlsx'
output_file = 'C:/Users/Administrator/Desktop/抖音热销商品排行榜20230511_transposed.xlsx'

# 打开Excel文件
try:
    wb = openpyxl.load_workbook(input_file)
    sheet = wb.active
except FileNotFoundError:
    print(f"文件 '{input_file}' 不存在，请检查文件路径是否正确！")
    exit()
except InvalidFileException:
    print(f"文件 '{input_file}' 无法打开，请检查文件格式是否正确！")
    exit()

# 新建Excel文件并获取第一个工作表
new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

# 遍历第一个工作表的第一列，并将数据填入新表格
max_row = sheet.max_row
current_row = 1
for i in range(1, max_row+1):
    cell = sheet.cell(row=i, column=1)
    value = cell.value
    data_type = cell.data_type
    if data_type == 's':
        # 文本类型
        new_sheet.cell(row=current_row, column=(i-1)%8+1, value=value)
        print(f"发现文本: {value}")
    elif data_type == 'n' and cell.is_date:
        # 日期类型
        new_sheet.cell(row=current_row, column=(i-1)%8+1, value=value.strftime('%Y-%m-%d'))
    elif data_type == 'n' and cell.number_format == '0.00%':
        # 百分比类型
        new_sheet.cell(row=current_row, column=(i-1)%8+1, value=value)
    elif data_type == 'n':
        # 数字类型
        new_sheet.cell(row=current_row, column=(i-1)%8+1, value=value)
    elif data_type == 'b':
        # 布尔类型
        new_sheet.cell(row=current_row, column=(i-1)%8+1, value=value)
    elif data_type == 'e':
        # 错误类型
        new_sheet.cell(row=current_row, column=(i-1)%8+1, value=value)
    elif data_type == 'i':
        # 图片类型
        img = Image(cell.value)
        img.width = 250
        img.height = 250
        new_sheet.column_dimensions[get_column_letter((i-1)%8+1)].width = 15
        new_sheet.row_dimensions[current_row].height = 70
        new_sheet.add_image(img, f'{get_column_letter((i-1)%8+1)}{current_row}')
        print("发现图片")
    if i % 8 == 0:
        current_row += 1

# 设置新表格中的单元格对齐方式
for row in new_sheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


# 保存新表格
try:
    new_wb.save(output_file)
    print(f"表格转换完成，已保存为文件: {output_file}")
except PermissionError:
    print(f"文件 '{output_file}' 无法保存，请检查文件是否被其他程序占用或是否具有写入权限！")
    
